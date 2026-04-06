"""
SNS's Study Plan — Google Sheets Builder
─────────────────────────────────────────
1.  pip install openpyxl
2.  python sns_study_plan_builder.py
3.  Go to sheets.google.com → File → Import → Upload the .xlsx → Replace spreadsheet
4.  Extensions → Apps Script → paste sns_apps_script.js → Save → Run setup()
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# ── Palette ───────────────────────────────────────────────────────────────────
W = "FFFFFFFF"
BLK = "FF212121"
HDR1 = "FF0D47A1"
HDR2 = "FF1B5E20"
HDR3 = "FF4A148C"
HDR4 = "FF006064"
HDR5 = "FF37474F"
DGREY = "FF455A64"
LGREY = "FFF5F5F5"
STRIPE = "FFF0F4FF"
YELLOW = "FFFFF9C4"
LT_BLU = "FFE3F2FD"
LT_GRN = "FFC8E6C9"
LT_RED = "FFFFCDD2"
LT_ORG = "FFFFE0B2"
MED_RED = "FFEF9A9A"
INIT_C = "FFBBDEFB"
PROG_C = "FFFFF9C4"
COMP_C = "FFE8F5E9"
DONE_C = "FFA5D6A7"
ARCH_C = "FFECEFF1"


# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(h):
    return PatternFill("solid", fgColor=h)


def fnt(bold=False, color=BLK, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")


def bdr(t="thin"):
    s = Side(style=t)
    return Border(left=s, right=s, top=s, bottom=s)


def aln(h="center", wrap=True, v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def set_cell(
    ws,
    row,
    col,
    value="",
    bg=W,
    bold=False,
    color=BLK,
    size=10,
    h="center",
    wrap=True,
    border=True,
    italic=False,
    num_fmt=None,
):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = Font(bold=bold, color=color, size=size, italic=italic, name="Arial")
    c.alignment = aln(h, wrap)
    if border:
        c.border = bdr()
    if num_fmt:
        c.number_format = num_fmt
    return c


def banner(ws, text, bg, cols, row=1, fg=W, sz=14, height=36):
    last = get_column_letter(cols)
    ws.merge_cells(f"A{row}:{last}{row}")
    set_cell(ws, row, 1, text, bg, bold=True, color=fg, size=sz)
    ws.row_dimensions[row].height = height


def sub_banner(ws, text, bg, cols, row, fg=W, sz=10, height=16):
    last = get_column_letter(cols)
    ws.merge_cells(f"A{row}:{last}{row}")
    set_cell(ws, row, 1, text, bg, bold=False, color=fg, size=sz, italic=True)
    ws.row_dimensions[row].height = height


def header_row(ws, row, cols_spec, bg, fg=W, height=36):
    """cols_spec = list of (label, width)"""
    for c, (label, width) in enumerate(cols_spec, 1):
        cell = set_cell(ws, row, c, label, bg, bold=True, color=fg, size=10)
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[row].height = height


def data_rows(ws, start, end, ncols, even=W, odd=LGREY, height=24):
    for r in range(start, end + 1):
        bg = even if r % 2 == 0 else odd
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill(bg)
            cell.font = fnt(size=10)
            cell.border = bdr()
            cell.alignment = aln()
        ws.row_dimensions[r].height = height


def add_dv(ws, formula, rows_range):
    dv = DataValidation(
        type="list", formula1=formula, allow_blank=True, showDropDown=False
    )
    ws.add_data_validation(dv)
    dv.add(rows_range)
    return dv


def cf_row(ws, rng, formula, bg, bold=False, color=BLK):
    ws.conditional_formatting.add(
        rng,
        FormulaRule(
            formula=[formula],
            fill=fill(bg),
            font=Font(bold=bold, color=color, name="Arial", size=10),
        ),
    )


wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# 1. CONFIG (hidden)
# ══════════════════════════════════════════════════════════════════════════════
cfg = wb.active
cfg.title = "⚙️ Config"
cfg.sheet_state = "hidden"

cfg_cols = {
    "A": (
        "Resource Types",
        [
            "YouTube Video",
            "Article/Blog",
            "PDF/Book",
            "Online Course",
            "Lecture Notes",
            "Problem Set",
            "Podcast",
            "Other",
        ],
    ),
    "B": (
        "Status Options",
        ["Not Started", "Initiated", "In Progress", "Completed", "On Hold"],
    ),
    "C": ("Done Options", ["✅ Done", "⏳ Pending", "❌ Skipped"]),
    "D": (
        "Session Types",
        ["New Study", "Day-4 Revision", "Day-7 Revision", "Practice", "Review"],
    ),
    "E": (
        "Platforms",
        [
            "TryHackMe",
            "HackTheBox",
            "YouTube",
            "Udemy",
            "Coursera",
            "Book",
            "Blog",
            "Other",
        ],
    ),
    "F": ("Path Status", ["Not Started", "In Progress", "On Hold", "Completed"]),
    "G": ("Confidence", ["⭐ 1", "⭐⭐ 2", "⭐⭐⭐ 3", "⭐⭐⭐⭐ 4", "⭐⭐⭐⭐⭐ 5"]),
    "H": ("Left Off Label", ["📍 Where I Left Off"]),
}
for col, (heading, items) in cfg_cols.items():
    cfg[f"{col}1"] = heading
    cfg[f"{col}1"].font = fnt(bold=True, size=10)
    for i, item in enumerate(items, 2):
        cfg[f"{col}{i}"] = item

# ══════════════════════════════════════════════════════════════════════════════
# 2. LEARNING PATHS
# ══════════════════════════════════════════════════════════════════════════════
lp = wb.create_sheet("🛤️ Learning Paths")
lp.sheet_view.showGridLines = False
lp.freeze_panes = "A4"

NC_LP = 11
banner(lp, "🛤️  SNS'S LEARNING PATHS", HDR2, NC_LP, sz=15)
sub_banner(
    lp,
    (
        "Type directly to add a learning path. Path names auto-populate all dropdowns. "
        "Completed paths archive their tasks here as a collapsible group below this table."
    ),
    "FFE8F5E9",
    NC_LP,
    row=2,
    fg=HDR2,
)

lp_cols = [
    ("#", 4),
    ("Learning Path Name", 28),
    ("Description", 32),
    ("Platform /\nSource", 15),
    ("Total Topics\nPlanned", 13),
    ("Topics\nCompleted", 13),
    ("% Complete", 12),
    ("📍 Where I Left Off\n(Current Topic)", 30),
    ("Last Studied\nDate", 15),
    ("Status", 15),
    ("Notes / Next Steps", 30),
]
header_row(lp, 3, lp_cols, HDR2, height=38)

data_rows(lp, 4, 103, NC_LP, W, "FFF1F8E9")
for r in range(4, 104):
    lp.cell(row=r, column=1).value = r - 3
    lp.cell(row=r, column=7).value = f'=IFERROR(F{r}/E{r},"")'
    lp.cell(row=r, column=7).number_format = "0%"
    lp.cell(row=r, column=9).number_format = "DD-MMM-YY"
    for c in (2, 3, 8, 11):
        lp.cell(row=r, column=c).alignment = aln("left")
    add_dv(lp, "'⚙️ Config'!$E$2:$E$9", f"D{r}")
    add_dv(lp, "'⚙️ Config'!$F$2:$F$5", f"J{r}")

rng = "A4:K103"
cf_row(lp, rng, '$J4="Completed"', DONE_C, bold=True, color="FF1B5E20")
cf_row(lp, rng, '$J4="In Progress"', PROG_C)
cf_row(lp, rng, '$J4="On Hold"', LT_ORG)
cf_row(lp, rng, '$J4="Not Started"', "FFE0E0E0")

# Archive section placeholder (Apps Script will populate below row 110)
lp.merge_cells("A106:K106")
set_cell(
    lp,
    106,
    1,
    "📦  COMPLETED PATH ARCHIVES  —  Managed automatically by Apps Script",
    DGREY,
    bold=True,
    color=W,
    size=11,
)
lp.row_dimensions[106].height = 28

lp.merge_cells("A107:K107")
set_cell(
    lp,
    107,
    1,
    "When a Learning Path is marked Completed, its tasks are moved here "
    "as a collapsible group. Click the + / − on the left row numbers to expand.",
    ARCH_C,
    italic=True,
    color=DGREY,
    size=9,
    h="left",
)
lp.row_dimensions[107].height = 20

# ══════════════════════════════════════════════════════════════════════════════
# 3. STUDY LOG
# ══════════════════════════════════════════════════════════════════════════════
sl = wb.create_sheet("📚 Study Log")
sl.sheet_view.showGridLines = False
sl.freeze_panes = "A4"

NC_SL = 16
banner(sl, "📚  SNS'S STUDY PLAN — 1-4-7 Spaced Repetition Log", HDR1, NC_SL, sz=14)
sub_banner(
    sl,
    (
        "🔵 Initiated  🟡 In Progress  🟢 Completed  🔴 Day-4 Due  🟠 Day-7 Due  ✅ Mastered  "
        "| Sorts: Active → by Path then Start Date | Completed → bottom | "
        "Click date cells (E/F/G) to use date picker"
    ),
    DGREY,
    NC_SL,
    row=2,
    fg=W,
)

sl_cols = [
    ("#", 4),
    ("Learning\nPath", 20),
    ("Topic / Task", 32),
    ("Subject /\nTag", 15),
    ("Start\nDate", 13),
    ("Initiated\nDate", 13),
    ("Completed\nDate", 13),
    ("Status", 15),
    ("Resource Link / Notes", 32),
    ("Resource\nType", 14),
    ("Day-4\nDate", 12),
    ("Day-4\nDone? ✓", 12),
    ("Day-7\nDate", 12),
    ("Day-7\nDone? ✓", 12),
    ("Confidence\n(1–5 ⭐)", 13),
    ("Remarks", 22),
]
header_row(sl, 3, sl_cols, HDR1, height=38)

data_rows(sl, 4, 203, NC_SL, W, STRIPE)
for r in range(4, 204):
    sl.cell(row=r, column=1).value = r - 3
    # Day-4 = Completed + 3
    sl.cell(row=r, column=11).value = f'=IF(G{r}<>"",G{r}+3,"")'
    sl.cell(row=r, column=11).number_format = "DD-MMM-YY"
    # Day-7 = Completed + 6
    sl.cell(row=r, column=13).value = f'=IF(G{r}<>"",G{r}+6,"")'
    sl.cell(row=r, column=13).number_format = "DD-MMM-YY"
    for c in (5, 6, 7):
        sl.cell(row=r, column=c).number_format = "DD-MMM-YY"
    for c in (3, 4, 9, 16):
        sl.cell(row=r, column=c).alignment = aln("left")
    add_dv(sl, "'🛤️ Learning Paths'!$B$4:$B$103", f"B{r}")
    add_dv(sl, "'⚙️ Config'!$B$2:$B$6", f"H{r}")
    add_dv(sl, "'⚙️ Config'!$A$2:$A$9", f"J{r}")
    add_dv(sl, "'⚙️ Config'!$C$2:$C$4", f"L{r}")
    add_dv(sl, "'⚙️ Config'!$C$2:$C$4", f"N{r}")
    add_dv(sl, "'⚙️ Config'!$G$2:$G$7", f"O{r}")

rng = "A4:P203"
cf_row(sl, rng, 'AND($L4="✅ Done",$N4="✅ Done")', DONE_C, bold=True, color="FF1B5E20")
cf_row(sl, rng, 'AND($K4=TODAY(),$L4<>"✅ Done")', LT_RED, bold=True, color="FFB71C1C")
cf_row(sl, rng, 'AND($M4=TODAY(),$N4<>"✅ Done")', LT_ORG, bold=True, color="FFE65100")
cf_row(
    sl,
    rng,
    'AND($K4<TODAY(),$K4<>"",$L4<>"✅ Done")',
    MED_RED,
    bold=True,
    color="FFC62828",
)
cf_row(sl, rng, '$H4="Initiated"', INIT_C)
cf_row(sl, rng, '$H4="In Progress"', PROG_C)
cf_row(sl, rng, 'AND($H4="Completed",$L4<>"✅ Done")', COMP_C)
cf_row(sl, rng, '$H4="Completed"', "FFDCEDC8")

# ══════════════════════════════════════════════════════════════════════════════
# 4. TASKS BY CATEGORY
# ══════════════════════════════════════════════════════════════════════════════
tc = wb.create_sheet("📋 Tasks by Category")
tc.sheet_view.showGridLines = False

NC_TC = 9
banner(
    tc, "📋  TASKS BY CATEGORY  —  Auto-grouped by Learning Path", HDR3, NC_TC, sz=14
)
sub_banner(
    tc,
    (
        "This view is auto-generated by Apps Script. "
        "Click '🔄 Refresh View' to rebuild after adding tasks or paths."
    ),
    "FFE8EAF6",
    NC_TC,
    row=2,
    fg=HDR3,
)

# Refresh button placeholder row
tc.merge_cells("A3:I3")
set_cell(
    tc,
    3,
    1,
    "🔄  REFRESH VIEW  (Run rebuildCategoryView() in Apps Script "
    "or use the button after setup)",
    "FFD1C4E9",
    bold=True,
    color=HDR3,
    size=10,
)
tc.row_dimensions[3].height = 26

tc_col_widths = [4, 32, 18, 13, 16, 12, 12, 13, 24]
for c, w in enumerate(tc_col_widths, 1):
    tc.column_dimensions[get_column_letter(c)].width = w

# Placeholder content — Apps Script rebuilds this
tc.merge_cells("A5:I5")
set_cell(
    tc,
    5,
    1,
    "📌  No data yet. Add learning paths in 🛤️ Learning Paths, "
    "add tasks in 📚 Study Log, then click Refresh.",
    YELLOW,
    italic=True,
    color=DGREY,
    size=10,
    h="left",
)
tc.row_dimensions[5].height = 28

# ══════════════════════════════════════════════════════════════════════════════
# 5. TASK SCHEDULER
# ══════════════════════════════════════════════════════════════════════════════
sch = wb.create_sheet("📅 Task Scheduler")
sch.sheet_view.showGridLines = False

NC_SCH = 9
banner(sch, "📅  SNS'S TASK SCHEDULER", HDR4, NC_SCH, sz=14)
sub_banner(
    sch,
    (
        "Plan your study sessions. Only non-completed tasks appear in dropdowns. "
        "Today panel auto-suggests overdue/due revisions."
    ),
    TEAL_LT := "FFE0F2F1",
    NC_SCH,
    row=2,
    fg=HDR4,
)

sch_col_widths = [13, 10, 34, 20, 18, 14, 14, 10, 26]
for c, w in enumerate(sch_col_widths, 1):
    sch.column_dimensions[get_column_letter(c)].width = w

# ── TODAY panel ───────────────────────────────────────────────────────────────
sch.merge_cells("A4:I4")
set_cell(sch, 4, 1, "📌  TODAY", HDR4, bold=True, color=W, size=12)
sch.row_dimensions[4].height = 26

sch.merge_cells("A5:I5")
set_cell(
    sch,
    5,
    1,
    '=("📅  " & TEXT(TODAY(),"DDDD, DD MMMM YYYY"))',
    LT_BLU,
    bold=True,
    color=HDR1,
    size=11,
)
sch.row_dimensions[5].height = 24

sch.merge_cells("A6:I6")
set_cell(
    sch,
    6,
    1,
    "⚡  Auto-suggested revisions due today (Apps Script fills this)",
    "FFFFE8E1",
    italic=True,
    color="FFBF360C",
    size=9,
)
sch.row_dimensions[6].height = 18

today_hdr = [
    (
        "Date",
        "Day",
        "Task",
        "Learning Path",
        "Session Type",
        "Planned Mins",
        "Actual Mins",
        "Done?",
        "Notes",
    )
]
for c, h in enumerate(today_hdr[0], 1):
    set_cell(sch, 7, c, h, HDR4, bold=True, color=W, size=10)
sch.row_dimensions[7].height = 28

# Auto-suggestion rows (Apps Script fills rows 8-12)
for r in range(8, 13):
    for c in range(1, NC_SCH + 1):
        cell = sch.cell(row=r, column=c)
        cell.fill = fill("FFE1F5FE")
        cell.font = fnt(size=9, italic=True, color=DGREY)
        cell.border = bdr()
        cell.alignment = aln()
    sch.cell(row=r, column=1).value = "—  auto  —"
    sch.row_dimensions[r].height = 22

# Manual today slots
sch.merge_cells("A13:I13")
set_cell(
    sch,
    13,
    1,
    "✏️  Manual slots — add extra tasks for today",
    "FFF3E5F5",
    italic=True,
    color=HDR3,
    size=9,
)
sch.row_dimensions[13].height = 18

for r in range(14, 19):
    data_rows(sch, r, r, NC_SCH, W, LGREY)
    add_dv(sch, "'⚙️ Config'!$D$2:$D$6", f"E{r}")
    add_dv(sch, "'⚙️ Config'!$C$2:$C$4", f"H{r}")
    sch.row_dimensions[r].height = 24

# ── THIS WEEK panel ───────────────────────────────────────────────────────────
sch.merge_cells("A20:I20")
set_cell(sch, 20, 1, "📆  THIS WEEK", HDR4, bold=True, color=W, size=12)
sch.row_dimensions[20].height = 26

sch.merge_cells("A21:D21")
set_cell(
    sch,
    21,
    1,
    '=("Week of: " & TEXT(TODAY()-WEEKDAY(TODAY(),2)+1,"DD MMM") '
    '& "  –  " & TEXT(TODAY()-WEEKDAY(TODAY(),2)+7,"DD MMM YYYY"))',
    LT_BLU,
    bold=True,
    color=HDR1,
    size=10,
)
sch.row_dimensions[21].height = 22

week_hdr = [
    "Date",
    "Day",
    "Task (dropdown)",
    "Learning Path",
    "Session Type",
    "Planned\nMins",
    "Actual\nMins",
    "Done?",
    "Notes",
]
for c, h in enumerate(week_hdr, 1):
    set_cell(sch, 22, c, h, HDR4, bold=True, color=W, size=10)
sch.row_dimensions[22].height = 30

days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
for i, day in enumerate(days):
    r = 23 + i * 3  # 3 rows per day (1 label + 2 task slots)
    bg_day = LT_BLU if day in ("Saturday", "Sunday") else "FFE8EAF6"

    # Day label row — cols A and B written separately (no merge)
    for c in range(1, NC_SCH + 1):
        cell = sch.cell(row=r, column=c)
        cell.fill = fill(bg_day)
        cell.font = fnt(bold=True, color=HDR4, size=10)
        cell.border = bdr()
        cell.alignment = aln()
    sch.cell(
        row=r, column=1
    ).value = f'=TEXT(TODAY()-WEEKDAY(TODAY(),2)+{i + 1},"DD-MMM")'
    sch.cell(row=r, column=1).number_format = "DD-MMM-YY"
    sch.cell(row=r, column=2).value = day
    sch.row_dimensions[r].height = 22

    # Task slots
    for slot in range(1, 3):
        sr = r + slot
        slot_bg = "FFFFF8E1" if slot == 1 else W
        data_rows(sch, sr, sr, NC_SCH, slot_bg, LGREY)
        sch.cell(
            row=sr, column=1
        ).value = f'=TEXT(TODAY()-WEEKDAY(TODAY(),2)+{i + 1},"DD-MMM")'
        sch.cell(row=sr, column=1).number_format = "DD-MMM-YY"
        sch.cell(row=sr, column=1).fill = fill(slot_bg)
        sch.cell(row=sr, column=1).font = fnt(color=DGREY, size=9)
        add_dv(sch, "'⚙️ Config'!$D$2:$D$6", f"E{sr}")
        add_dv(sch, "'⚙️ Config'!$C$2:$C$4", f"H{sr}")
        sch.row_dimensions[sr].height = 22

# ── Scheduler CF ──────────────────────────────────────────────────────────────
sch_rng = "A23:I44"
cf_row(sch, sch_rng, '$H23="✅ Done"', "FFE0E0E0", color="FF9E9E9E")
cf_row(sch, sch_rng, '$E23="Day-4 Revision"', LT_RED, color="FFB71C1C")
cf_row(sch, sch_rng, '$E23="Day-7 Revision"', LT_ORG, color="FFE65100")

# ══════════════════════════════════════════════════════════════════════════════
# 6. PROGRESS & CHARTS
# ══════════════════════════════════════════════════════════════════════════════
pr = wb.create_sheet("📊 Progress & Charts")
pr.sheet_view.showGridLines = False

NC_PR = 8
banner(pr, "📊  SNS'S PROGRESS & CHARTS", HDR5, NC_PR, sz=14)
sub_banner(
    pr,
    (
        "Overall stats update live. Path cards + doughnut charts are "
        "built by Apps Script. Click '🔄 Rebuild Charts' after adding paths."
    ),
    "FFE8EAF6",
    NC_PR,
    row=2,
    fg=HDR5,
)

pr_col_widths = [22, 14, 14, 14, 14, 14, 14, 14]
for c, w in enumerate(pr_col_widths, 1):
    pr.column_dimensions[get_column_letter(c)].width = w

# Overall stats strip
pr.merge_cells("A4:H4")
set_cell(pr, 4, 1, "📈  OVERALL STATS", HDR5, bold=True, color=W, size=11)
pr.row_dimensions[4].height = 26

stat_labels = [
    ("Total Paths", "=COUNTA('🛤️ Learning Paths'!B4:B103)", LT_BLU),
    ("Active Paths", "=COUNTIF('🛤️ Learning Paths'!J4:J103,\"In Progress\")", PROG_C),
    ("Total Tasks", "=COUNTA('📚 Study Log'!C4:C203)", LT_BLU),
    ("Completed", "=COUNTIF('📚 Study Log'!H4:H203,\"Completed\")", LT_GRN),
    (
        "Mastered",
        "=COUNTIFS('📚 Study Log'!L4:L203,\"✅ Done\",'📚 Study Log'!N4:N203,\"✅ Done\")",
        DONE_C,
    ),
    (
        "Due Today",
        "=COUNTIFS('📚 Study Log'!K4:K203,TODAY(),'📚 Study Log'!L4:L203,\"<>✅ Done\")",
        LT_RED,
    ),
    (
        "Overdue",
        '=COUNTIFS(\'📚 Study Log\'!K4:K203,"<"&TODAY(),\'📚 Study Log\'!L4:L203,"<>✅ Done",\'📚 Study Log\'!K4:K203,"<>"")',
        MED_RED,
    ),
    ("On Hold", "=COUNTIF('📚 Study Log'!H4:H203,\"On Hold\")", LT_ORG),
]
for c, (label, formula, bg) in enumerate(stat_labels, 1):
    set_cell(pr, 5, c, label, DGREY, bold=True, color=W, size=9)
    set_cell(pr, 6, c, formula, bg, bold=True, size=13)
    pr.row_dimensions[5].height = 20
    pr.row_dimensions[6].height = 32

# Charts placeholder
pr.merge_cells("A8:H8")
set_cell(
    pr,
    8,
    1,
    "🔄  PATH CARDS & DOUGHNUT CHARTS  —  Click Rebuild Charts to generate",
    "FFD1C4E9",
    bold=True,
    color=HDR3,
    size=10,
)
pr.row_dimensions[8].height = 26

pr.merge_cells("A9:H9")
set_cell(
    pr,
    9,
    1,
    "One card + doughnut chart per learning path is auto-built by Apps Script. "
    "Charts update live once created. Run rebuildProgressCharts() from Extensions → Apps Script.",
    YELLOW,
    italic=True,
    color=DGREY,
    size=9,
    h="left",
)
pr.row_dimensions[9].height = 28

# ══════════════════════════════════════════════════════════════════════════════
# 7. RESOURCE LIBRARY
# ══════════════════════════════════════════════════════════════════════════════
rl = wb.create_sheet("🔗 Resource Library")
rl.sheet_view.showGridLines = False
rl.freeze_panes = "A4"

NC_RL = 8
banner(rl, "🔗  RESOURCE LIBRARY", HDR3, NC_RL, sz=14)
sub_banner(
    rl,
    "Catalog all resources. Link to a Learning Path to surface them in Tasks by Category.",
    "FFE8EAF6",
    NC_RL,
    row=2,
    fg=HDR3,
)

rl_cols = [
    ("#", 4),
    ("Resource Title", 36),
    ("Type", 16),
    ("URL / Location", 44),
    ("Learning Path", 22),
    ("Related Topic", 24),
    ("Quality ⭐", 12),
    ("Notes", 26),
]
header_row(rl, 3, rl_cols, HDR3, height=32)

data_rows(rl, 4, 203, NC_RL, W, "FFF3E5F5")
for r in range(4, 204):
    rl.cell(row=r, column=1).value = r - 3
    for c in (2, 4, 6, 8):
        rl.cell(row=r, column=c).alignment = aln("left")
    add_dv(rl, "'⚙️ Config'!$A$2:$A$9", f"C{r}")
    add_dv(rl, "'🛤️ Learning Paths'!$B$4:$B$103", f"E{r}")
    add_dv(rl, "'⚙️ Config'!$G$2:$G$7", f"G{r}")

# ══════════════════════════════════════════════════════════════════════════════
# 8. HOW TO USE
# ══════════════════════════════════════════════════════════════════════════════
ht = wb.create_sheet("❓ How To Use")
ht.sheet_view.showGridLines = False
ht.column_dimensions["A"].width = 90

banner(ht, "❓  HOW TO USE SNS'S STUDY PLAN", DGREY, 1, sz=14)

guide = [
    ("SETUP (do once)", HDR1, True),
    (
        "  1. Import this .xlsx into Google Sheets (File → Import → Replace spreadsheet).",
        None,
        False,
    ),
    (
        "  2. Go to Extensions → Apps Script → paste the contents of sns_apps_script.js → Save.",
        None,
        False,
    ),
    (
        "  3. Run the setup() function once — this creates named ranges and installs triggers.",
        None,
        False,
    ),
    ("  4. Reload the spreadsheet. You're ready!", None, False),
    ("", None, False),
    ("ADDING LEARNING PATHS", HDR2, True),
    (
        "  Open 🛤️ Learning Paths. Type the path name in col B — it instantly appears in all dropdowns.",
        None,
        False,
    ),
    (
        "  Fill in Description, Platform, Total Topics Planned, Where I Left Off, and Status.",
        None,
        False,
    ),
    (
        "  Update '📍 Where I Left Off' (col H) as you progress through the path.",
        None,
        False,
    ),
    ("", None, False),
    ("LOGGING A TASK (Study Log)", HDR1, True),
    (
        "  Select the Learning Path from dropdown (col B). Enter Topic, Tag, and click col E for Start Date.",
        None,
        False,
    ),
    (
        "  Click col F (Initiated Date) when you begin. Click col G (Completed Date) when done.",
        None,
        False,
    ),
    (
        "  Status auto-sets: Not Started → Initiated → In Progress → Completed.",
        None,
        False,
    ),
    (
        "  Day-4 and Day-7 dates calculate automatically from the Completed Date.",
        None,
        False,
    ),
    (
        "  Row sorting: active tasks sorted by Path → Start Date. Completed tasks sink to bottom.",
        None,
        False,
    ),
    ("", None, False),
    ("DOING REVISIONS", HDR4, True),
    (
        "  🔴 Red row = Day-4 due today. Do a 15–20 min review → mark ✅ Done in col L.",
        None,
        False,
    ),
    (
        "  🟠 Orange row = Day-7 due today. Do active recall → mark ✅ Done in col N.",
        None,
        False,
    ),
    ("  ✅ Pale green row = both revisions done — topic fully mastered!", None, False),
    ("", None, False),
    ("TASK SCHEDULER", HDR4, True),
    (
        "  Today panel (top): auto-suggests revisions due today. 5 manual slots for extra tasks.",
        None,
        False,
    ),
    (
        "  This Week panel: plan tasks per day. Dropdown only shows non-completed tasks.",
        None,
        False,
    ),
    ("  Selecting a task auto-fills its Learning Path.", None, False),
    ("", None, False),
    ("COMPLETING A LEARNING PATH", HDR2, True),
    ("  Set Status = Completed in 🛤️ Learning Paths col J.", None, False),
    (
        "  Apps Script automatically moves all its tasks from Study Log into an",
        None,
        False,
    ),
    (
        "  archived collapsible group below the main paths table in Learning Paths sheet.",
        None,
        False,
    ),
    (
        "  Click +/− on the left row numbers to expand/collapse the archive.",
        None,
        False,
    ),
    ("", None, False),
    ("CHARTS & PROGRESS", HDR5, True),
    ("  Open 📊 Progress & Charts. Overall stats update live.", None, False),
    (
        "  Click 'Rebuild Charts' button to regenerate per-path doughnut charts.",
        None,
        False,
    ),
    ("  Run this after adding new learning paths.", None, False),
    ("", None, False),
    ("RESOURCES", HDR3, True),
    (
        "  Add resources in 🔗 Resource Library and link them to a Learning Path.",
        None,
        False,
    ),
    (
        "  They appear automatically in that path's block in 📋 Tasks by Category.",
        None,
        False,
    ),
    ("", None, False),
    ("COLUMN REFERENCE — Study Log", DGREY, True),
    (
        "  A=#  B=Path  C=Topic  D=Tag  E=Start  F=Initiated  G=Completed  H=Status",
        None,
        False,
    ),
    ("  I=Resource Link  J=Resource Type  K=Day-4 Date  L=Day-4 Done", None, False),
    ("  M=Day-7 Date  N=Day-7 Done  O=Confidence  P=Remarks", None, False),
]

for i, (text, color, bold) in enumerate(guide, 2):
    ht.merge_cells(f"A{i}:A{i}")
    cell = ht.cell(row=i, column=1, value=text)
    cell.font = Font(
        bold=bold,
        size=11 if bold else 10,
        name="Arial",
        color=W if (bold and color) else BLK,
    )
    cell.fill = fill(color) if (bold and color) else fill(W if i % 2 == 0 else LGREY)
    cell.alignment = aln("left")
    ht.row_dimensions[i].height = 20 if bold else 18

# ── Sheet order ───────────────────────────────────────────────────────────────
order = [
    "🛤️ Learning Paths",
    "📚 Study Log",
    "📋 Tasks by Category",
    "📅 Task Scheduler",
    "📊 Progress & Charts",
    "🔗 Resource Library",
    "❓ How To Use",
    "⚙️ Config",
]
for i, name in enumerate(order):
    idx = wb.sheetnames.index(name)
    wb.move_sheet(name, offset=i - idx)

wb.save("SNS_Study_Plan.xlsx")
print("✅  SNS_Study_Plan.xlsx created!")
print()
print("Next steps:")
print("  1. Go to sheets.google.com")
print("  2. File → Import → Upload → SNS_Study_Plan.xlsx → Replace spreadsheet")
print("  3. Extensions → Apps Script → paste sns_apps_script.js → Save")
print("  4. Run setup() once")
print("  5. Reload the sheet — all triggers and named ranges are live!")
